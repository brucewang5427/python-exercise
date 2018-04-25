import openpyxl,csv,os
os.makedirs('exceltocsv',exist_ok=True)
for file in os.listdir('.'):
    if not file.endswith('.xlsx'):
        continue
    wb=openpyxl.load_workbook(file)
    for sheetname in wb.get_sheet_names():
        sheet=wb.get_sheet_by_name(sheetname)
        csvoutput=open(os.path.join('exceltocsv',sheetname+'.csv'),'w',newline='')
        csvwriter=csv.writer(csvoutput)
        for rowNum in range(1,sheet.max_row+1):
            rows=[]
            for colNum in range(1,sheet.max_column+1):
                rows.append(sheet.cell(row=rowNum,column=colNum).value)
            csvwriter.writerow(rows)
        csvoutput.close()